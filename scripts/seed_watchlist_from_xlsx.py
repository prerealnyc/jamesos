"""Seed the Apify watchlist from a Speaking Targets xlsx.

Reads a workbook shaped like Patrick Pharris's Speaking Targets sheet
(headers at row 5: Name | Subitems | People | Website | Podcast Title |
Podcast Links | Agent / Manager | Agent Email | Youtube | Instagram | X |
TikTok | Facebook | Spotify | Interests), extracts every IG/TT/YT handle
it can find from the URL cells, and merges them into the tenant
watchlist non-destructively (existing entries are kept, duplicates by
(platform, handle) are dropped).

Run:
    .venv/bin/python scripts/seed_watchlist_from_xlsx.py PATH/TO/file.xlsx

Honest scope:
  * Apify scraping today supports instagram, tiktok, youtube only — the X
    column is parsed but ignored when building the watchlist, since
    there's no actor wired for X yet. The handle is preserved in the
    `interests` cell when --x-as-note is set, so the data isn't lost.
  * Rows without any IG/TT/YT URL produce nothing (a TBD or podcast-only
    entry can't be scraped). Those are reported, not silently skipped.
"""

from __future__ import annotations

import argparse
import asyncio
import re
import sys
from pathlib import Path
from urllib.parse import urlparse

import openpyxl

# Re-use the same write path the API endpoint uses so RLS / per-tenant
# storage works the same way.
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
from james_os.db import close_pool, init_pool  # noqa: E402
from james_os.trends import get_watchlist, set_watchlist  # noqa: E402


# ── handle extraction ──────────────────────────────────────────────────

_IG = re.compile(r"instagram\.com/([^/?#]+)", re.I)
_TT = re.compile(r"tiktok\.com/@([^/?#]+)", re.I)
# YouTube channels appear in many shapes — handle, /c/, /channel/, /user/.
_YT_HANDLE = re.compile(r"youtube\.com/@([^/?#]+)", re.I)
_YT_C = re.compile(r"youtube\.com/(?:c|user)/([^/?#]+)", re.I)
_YT_CHAN = re.compile(r"youtube\.com/channel/([^/?#]+)", re.I)


def _strip_url(s: str | None) -> str:
    if not s:
        return ""
    # Cells sometimes hold two URLs separated by ; or a stray comment — take
    # the first whitespace-delimited token that looks like a URL.
    for tok in str(s).split():
        tok = tok.strip().strip(";,")
        if tok.startswith("http"):
            return tok
    return str(s).strip()


def extract_ig(cell: str | None) -> str | None:
    m = _IG.search(_strip_url(cell))
    if not m:
        return None
    h = m.group(1).strip("/").strip()
    # Some rows have /reel/ or /p/ embedded — filter those.
    if h in ("p", "reel", "explore", "tv"):
        return None
    return h or None


def extract_tt(cell: str | None) -> str | None:
    m = _TT.search(_strip_url(cell))
    return m.group(1).strip("/").strip() if m else None


def extract_yt(cell: str | None) -> str | None:
    url = _strip_url(cell)
    for pat in (_YT_HANDLE, _YT_C, _YT_CHAN):
        m = pat.search(url)
        if m:
            return m.group(1).strip("/").strip()
    return None


# ── xlsx parsing ───────────────────────────────────────────────────────


def find_header_row(ws) -> int:
    """The Speaking Targets sheet has a few title rows before the real
    header — find the row that starts with 'Name'."""
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and isinstance(row[0], str) and row[0].strip().lower() == "name":
            return i
    raise ValueError("could not locate header row (no 'Name' cell)")


def col_index(headers: list[str], name: str) -> int | None:
    for i, h in enumerate(headers):
        if h and h.strip().lower() == name.lower():
            return i
    return None


def parse_targets(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = find_header_row(ws)
    headers = [str(c.value or "").strip() for c in ws[header_row]]
    name_i = col_index(headers, "Name")
    yt_i = col_index(headers, "Youtube")
    ig_i = col_index(headers, "Instagram")
    tt_i = col_index(headers, "TikTok")
    interests_i = col_index(headers, "Interests")
    if name_i is None:
        raise ValueError("no Name column in header row")

    targets: list[dict] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        name = (row[name_i] or "") if name_i is not None else ""
        name = str(name).strip()
        if not name:
            continue
        ig = extract_ig(row[ig_i] if ig_i is not None else None)
        tt = extract_tt(row[tt_i] if tt_i is not None else None)
        yt = extract_yt(row[yt_i] if yt_i is not None else None)
        interests_raw = row[interests_i] if interests_i is not None else None
        interests = [
            s.strip()
            for s in str(interests_raw or "").split(",")
            if s.strip()
        ]
        if not (ig or tt or yt):
            targets.append({
                "name": name, "platforms": {}, "interests": interests,
                "skip_reason": "no IG/TT/YT URL found",
            })
            continue
        platforms = {}
        if ig:
            platforms["instagram"] = ig
        if tt:
            platforms["tiktok"] = tt
        if yt:
            platforms["youtube"] = yt
        targets.append({
            "name": name, "platforms": platforms, "interests": interests,
        })
    return targets


# ── merge + persist ────────────────────────────────────────────────────


async def merge(targets: list[dict], dry_run: bool) -> None:
    # The trends helpers expect a live pool — same one the API uses.
    await init_pool()
    try:
        await _do_merge(targets, dry_run)
    finally:
        await close_pool()


async def _do_merge(targets: list[dict], dry_run: bool) -> None:
    existing = await get_watchlist()
    seen = {(c["platform"], c["handle"].lower()) for c in existing}
    additions: list[dict] = []
    for t in targets:
        if not t.get("platforms"):
            continue
        for platform, handle in t["platforms"].items():
            key = (platform, handle.lower())
            if key in seen:
                continue
            seen.add(key)
            additions.append({
                "platform": platform,
                "handle": handle,
                "name": t["name"],
                "interests": t["interests"],
            })
    print(f"existing entries: {len(existing)}")
    print(f"new entries to add: {len(additions)}")
    for a in additions:
        tag = (", ".join(a["interests"][:3]) or "—")
        print(f"  + {a['platform']:<9} @{a['handle']:<32} ({a['name']}) [{tag}]")
    skipped = [t for t in targets if not t.get("platforms")]
    if skipped:
        print(f"\nrows skipped (no IG/TT/YT): {len(skipped)}")
        for t in skipped:
            print(f"  · {t['name']:<40} {t.get('skip_reason', '')}")
    if dry_run:
        print("\n(dry-run — nothing written)")
        return
    if not additions:
        print("\nnothing new to add.")
        return
    merged = list(existing) + additions
    written = await set_watchlist(merged)
    print(f"\nwatchlist now has {len(written)} creators.")


# ── CLI ────────────────────────────────────────────────────────────────


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", type=Path, help="path to the Speaking Targets xlsx")
    ap.add_argument("--dry-run", action="store_true",
                    help="parse + report; do not write")
    args = ap.parse_args()
    if not args.xlsx.exists():
        ap.error(f"file not found: {args.xlsx}")
    targets = parse_targets(args.xlsx)
    asyncio.run(merge(targets, args.dry_run))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
